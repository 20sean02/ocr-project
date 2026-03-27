#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import io
import csv
import uuid
import hashlib
from datetime import datetime

from flask import Flask, request, render_template, send_file, jsonify

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import extract_subjects_batch_full_withEN as ocrmod

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB

# In-memory storage
RESULTS = {}          # token -> list of row dicts (editable)
IMAGE_CACHE = {}      # file content hash -> row dict (skip duplicates)
EXPORT_LOG = []       # list of {"filename", "time", "rows", "format"}

# Server-side export directory
EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "csv_exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

IMAGES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images")
os.makedirs(IMAGES_DIR, exist_ok=True)

HEADERS = ["工作項目編號", "來文機關", "來文字號", "收文文號", "收文日期", "事由"]
FIELD_MAP = [
    ("工作項目編號", "工作項目編號"),
    ("來文機關", "來文機關"),
    ("來文字號", "來文字號"),
    ("收文文號", "收文文號"),
    ("收文日期", "收文日期"),
    ("事由", "事由"),
]


def _hash_file(file_obj) -> str:
    data = file_obj.read()
    file_obj.seek(0)
    return hashlib.sha256(data).hexdigest()


def _rows_to_csv(rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for r in rows:
        writer.writerow([r.get(key, "") or "" for key, _ in FIELD_MAP])
    return buf.getvalue()


XLSX_HEADERS = ["工作項目編號", "來文機關", "來文字號", "收文文號", "收文日期", "事由"]
XLSX_FIELD_MAP = [
    ("工作項目編號", "工作項目編號"),
    ("來文機關", "來文機關"),
    ("來文字號", "來文字號"),
    ("收文文號", "收文文號"),
    ("收文日期", "收文日期"),
    ("事由", "事由"),
]


def _rows_to_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "辨識結果"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    for col, name in enumerate(XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for row_idx, r in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(XLSX_FIELD_MAP, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, "") or "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "事由"))

    col_widths = [16, 16, 16, 14, 12, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Export helpers ──────────────────────────────────────────────

def _save_export(rows, token, fmt, custom_name=None):
    ext = {"csv": ".csv", "xlsx": ".xlsx"}.get(fmt, ".csv")
    if custom_name:
        # Sanitize and ensure correct extension
        custom_name = os.path.basename(custom_name).strip()
        if not custom_name.endswith(ext):
            custom_name = os.path.splitext(custom_name)[0] + ext
        filename = custom_name
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{ts}_{token}{ext}"
    path = os.path.join(EXPORT_DIR, filename)

    if fmt == "xlsx":
        buf = _rows_to_xlsx(rows)
        with open(path, "wb") as f:
            f.write(buf.getvalue())
    else:
        csv_data = _rows_to_csv(rows)
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_data)

    EXPORT_LOG.append({
        "filename": filename,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rows": len(rows),
        "format": fmt.upper(),
    })
    if len(EXPORT_LOG) > 100:
        EXPORT_LOG[:] = EXPORT_LOG[-100:]

    return filename


def _sort_and_renumber(rows):
    """Sort rows by 收文日期 ascending and re-assign 工作項目編號."""
    rows.sort(key=lambda r: (
        r.get("_sort_key") or ocrmod.parse_roc_date_to_sort_key(r.get("收文日期", ""))
    ))
    _renumber(rows)


def _renumber(rows):
    """Re-assign 工作項目編號 without changing order."""
    for idx, r in enumerate(rows, start=1):
        r["工作項目編號"] = ocrmod.format_work_item_id(idx, r.get("收文日期", ""))


def _read_csv_file(path):
    """Read a CSV export file and return list of row dicts."""
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({
                "工作項目編號": r.get("工作項目編號", ""),
                "original_filename": r.get("原始檔名", ""),
                "來文機關": r.get("來文機關", "N/A"),
                "來文字號": r.get("來文字號", "N/A"),
                "收文文號": r.get("收文文號", ""),
                "收文日期": r.get("收文日期", ""),
                "事由": r.get("事由", ""),
                "status": r.get("status", ""),
            })
    return rows


def _read_xlsx_file(path):
    """Read an XLSX export file and return list of row dicts."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        wb.close()
        return []
    headers = [str(h).strip() if h else "" for h in headers]
    rows = []
    for vals in rows_iter:
        r = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, vals)}
        rows.append({
            "工作項目編號": r.get("工作項目編號", ""),
            "original_filename": r.get("原始檔名", ""),
            "來文機關": r.get("來文機關", "N/A"),
            "來文字號": r.get("來文字號", "N/A"),
            "收文文號": r.get("收文文號", ""),
            "收文日期": r.get("收文日期", ""),
            "事由": r.get("事由", ""),
            "status": r.get("status", ""),
        })
    wb.close()
    return rows


# ── Routes ─────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/images/<path:filename>", methods=["GET"])
def serve_image(filename):
    safe = os.path.basename(filename)
    img_path = os.path.join(IMAGES_DIR, safe)
    if not os.path.isfile(img_path):
        return "Not found", 404
    return send_file(img_path)


@app.route("/upload_one", methods=["POST"])
def upload_one():
    """Process a single image. The frontend calls this once per file so the
    user can cancel between images."""
    ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
    f = request.files.get("file")
    if not f or f.filename == "":
        return jsonify({"error": "缺少圖片"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXT:
        return jsonify({"error": f"不支援的檔案格式：{ext}，請上傳圖片檔"}), 400

    token = request.form.get("token", "").strip()
    if not token or token not in RESULTS:
        token = uuid.uuid4().hex[:12]
        RESULTS[token] = []

    rows = RESULTS[token]
    existing_hashes = {r.get("_file_hash") for r in rows if r.get("_file_hash")}

    orig = os.path.basename(f.filename) or "upload.jpg"
    file_hash = _hash_file(f)

    # Duplicate check
    if file_hash in existing_hashes:
        return jsonify({"token": token, "skipped": True})

    # Cache hit
    if file_hash in IMAGE_CACHE:
        cached = IMAGE_CACHE[file_hash].copy()
        cached["original_filename"] = orig
        cached["_file_hash"] = file_hash
        rows.append(cached)
        _sort_and_renumber(rows)
        return jsonify({"token": token, "skipped": True, "cached": True,
                        "row": _safe_row(cached)})

    # Save image to images/ folder for later viewing
    save_name = f"{file_hash[:10]}_{orig}"
    save_path = os.path.join(IMAGES_DIR, save_name)
    f.save(save_path)

    # Process with OCR
    try:
        subject, issue_date, laiwen_dept, doc_no, from_agency = ocrmod.process_one(save_path)
        status = "ok" if (subject and from_agency != "N/A" and doc_no and issue_date) else "partial"
        row = {
            "original_filename": orig,
            "來文機關": from_agency or "N/A",
            "來文字號": laiwen_dept or "N/A",
            "收文文號": doc_no or "",
            "收文日期": issue_date or "",
            "事由": subject or "",
            "status": status,
            "_sort_key": ocrmod.parse_roc_date_to_sort_key(issue_date),
            "_file_hash": file_hash,
            "_image_file": save_name,
        }
    except Exception as e:
        row = {
            "original_filename": orig,
            "來文機關": "N/A", "來文字號": "N/A",
            "收文文號": "", "收文日期": "", "事由": "",
            "status": f"error: {e}",
            "_sort_key": (1, 0, 0, 0),
            "_file_hash": file_hash,
            "_image_file": save_name,
        }

    rows.append(row)
    IMAGE_CACHE[file_hash] = row.copy()
    _sort_and_renumber(rows)

    if len(RESULTS) > 20:
        oldest = list(RESULTS.keys())[0]
        del RESULTS[oldest]
    if len(IMAGE_CACHE) > 500:
        keys = list(IMAGE_CACHE.keys())
        for k in keys[:250]:
            del IMAGE_CACHE[k]

    return jsonify({"token": token, "skipped": False, "row": _safe_row(row)})


def _cleanup_images(rows):
    """Delete image files associated with rows."""
    for r in rows:
        img = r.get("_image_file")
        if img:
            path = os.path.join(IMAGES_DIR, img)
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except OSError:
                pass


def _safe_row(r):
    """Return a row dict safe for JSON (no tuple sort keys)."""
    return {
        "工作項目編號": r.get("工作項目編號", ""),
        "original_filename": r.get("original_filename", ""),
        "來文機關": r.get("來文機關", "N/A"),
        "來文字號": r.get("來文字號", "N/A"),
        "收文文號": r.get("收文文號", ""),
        "收文日期": r.get("收文日期", ""),
        "事由": r.get("事由", ""),
        "status": r.get("status", ""),
        "_file_hash": r.get("_file_hash", ""),
        "_image_file": r.get("_image_file", ""),
    }


@app.route("/get_rows/<token>", methods=["GET"])
def get_rows(token):
    """Return all current rows for a token (used after processing to get
    renumbered results)."""
    rows = RESULTS.get(token)
    if rows is None:
        return jsonify({"error": "token not found"}), 404
    safe_rows = [_safe_row(r) for r in rows]
    ok_count = sum(1 for r in rows if r.get("status") == "ok")
    partial_count = sum(1 for r in rows if r.get("status") == "partial")
    error_count = sum(1 for r in rows if r.get("status", "").startswith("error"))
    return jsonify({
        "token": token, "rows": safe_rows,
        "ok_count": ok_count, "partial_count": partial_count,
        "error_count": error_count,
    })


def _get_edited_rows(token):
    """Extract edited rows and optional filename from request, update caches."""
    stored_rows = RESULTS.get(token)
    if stored_rows is None:
        return None, None

    data = request.get_json(silent=True)
    custom_name = None

    # Accept both {rows: [...], filename: "..."} and plain [...]
    if data and isinstance(data, dict):
        edited = data.get("rows")
        custom_name = data.get("filename")
    elif data and isinstance(data, list):
        edited = data
    else:
        edited = None

    if edited and isinstance(edited, list):
        rows = edited
        for edited_row, stored_row in zip(edited, stored_rows):
            file_hash = edited_row.get("_file_hash") or stored_row.get("_file_hash")
            if file_hash and file_hash in IMAGE_CACHE:
                for field in ("來文機關", "來文字號", "收文文號", "收文日期", "事由"):
                    if field in edited_row:
                        IMAGE_CACHE[file_hash][field] = edited_row[field]
        RESULTS[token] = edited
    else:
        rows = stored_rows
    return rows, custom_name


@app.route("/download/<token>/csv", methods=["POST"])
def download_csv(token):
    rows, custom_name = _get_edited_rows(token)
    if rows is None:
        return "結果已過期，請重新上傳辨識", 404

    csv_data = _rows_to_csv(rows)
    _save_export(rows, token, "csv", custom_name)
    _cleanup_images(rows)

    dl_name = custom_name or "subject_output.csv"
    if not dl_name.endswith(".csv"):
        dl_name = os.path.splitext(dl_name)[0] + ".csv"
    buf = io.BytesIO(csv_data.encode("utf-8-sig"))
    return send_file(buf, as_attachment=True, download_name=dl_name, mimetype="text/csv")


@app.route("/download/<token>/xlsx", methods=["POST"])
def download_xlsx(token):
    rows, custom_name = _get_edited_rows(token)
    if rows is None:
        return "結果已過期，請重新上傳辨識", 404

    _save_export(rows, token, "xlsx", custom_name)
    _cleanup_images(rows)
    dl_name = custom_name or "subject_output.xlsx"
    if not dl_name.endswith(".xlsx"):
        dl_name = os.path.splitext(dl_name)[0] + ".xlsx"
    buf = _rows_to_xlsx(rows)

    return send_file(
        buf, as_attachment=True,
        download_name=dl_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



@app.route("/download/<token>/append", methods=["POST"])
def download_append(token):
    """Append current data to an existing CSV export file."""
    stored_rows = RESULTS.get(token)
    if stored_rows is None:
        return "結果已過期，請重新上傳辨識", 404

    data = request.get_json(silent=True)
    if not data or not data.get("target"):
        return "Missing target filename", 400

    target_name = data["target"]
    edited_rows = data.get("rows", stored_rows)

    if ".." in target_name or "/" in target_name:
        return "Invalid filename", 400
    target_path = os.path.join(EXPORT_DIR, target_name)
    if not os.path.isfile(target_path):
        return "目標檔案不存在", 404

    # Read existing CSV
    existing = _read_csv_file(target_path)

    # Merge: append new rows, skip duplicates by 收文文號
    existing_doc_nos = {r.get("收文文號") for r in existing if r.get("收文文號")}
    duplicates = []
    added = 0
    for r in edited_rows:
        doc_no = r.get("收文文號", "")
        if doc_no and doc_no in existing_doc_nos:
            duplicates.append(doc_no)
            continue
        existing.append(r)
        added += 1
        if doc_no:
            existing_doc_nos.add(doc_no)

    # If all rows were duplicates, don't write and inform user
    if added == 0:
        return jsonify({
            "ok": False,
            "all_duplicates": True,
            "duplicates": duplicates,
            "total_rows": len(existing),
            "filename": target_name,
        })

    # Sort by date and re-number
    _sort_and_renumber(existing)

    # Write back
    csv_data = _rows_to_csv(existing)
    with open(target_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write(csv_data)

    _cleanup_images(edited_rows)

    return jsonify({
        "ok": True,
        "all_duplicates": False,
        "duplicates": duplicates,
        "total_rows": len(existing),
        "filename": target_name,
    })


@app.route("/exports", methods=["GET"])
def list_exports():
    files = []
    for name in sorted(os.listdir(EXPORT_DIR), reverse=True):
        if name.endswith((".csv", ".xlsx")):
            path = os.path.join(EXPORT_DIR, name)
            stat = os.stat(path)
            fmt = "XLSX" if name.endswith(".xlsx") else "CSV"
            files.append({
                "filename": name,
                "size": stat.st_size,
                "time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "format": fmt,
            })
    return jsonify(files)


@app.route("/exports/<filename>", methods=["GET"])
def download_export(filename):
    if ".." in filename or "/" in filename:
        return "Invalid filename", 400
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.isfile(path):
        return "檔案不存在", 404
    mimes = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".csv": "text/csv",
    }
    ext = os.path.splitext(filename)[1]
    mime = mimes.get(ext, "application/octet-stream")
    return send_file(path, as_attachment=True, download_name=filename, mimetype=mime)


@app.route("/exports/<filename>/preview", methods=["GET"])
def preview_export(filename):
    if ".." in filename or "/" in filename:
        return "Invalid filename", 400
    if not filename.endswith(".csv"):
        return "只支援預覽 CSV 檔案", 400
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.isfile(path):
        return "檔案不存在", 404
    rows = _read_csv_file(path)
    return jsonify({"headers": HEADERS, "rows": rows})


@app.route("/exports/<filename>/save", methods=["POST"])
def save_export(filename):
    """Save edited rows back to a CSV file."""
    if ".." in filename or "/" in filename:
        return "Invalid filename", 400
    if not filename.endswith(".csv"):
        return "只支援編輯 CSV 檔案", 400
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.isfile(path):
        return "檔案不存在", 404

    rows = request.get_json(silent=True)
    if not rows or not isinstance(rows, list):
        return "無效的資料", 400

    _renumber(rows)
    csv_data = _rows_to_csv(rows)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        f.write(csv_data)

    return jsonify({"ok": True, "rows": len(rows)})


@app.route("/exports/<filename>/download/<fmt>", methods=["GET"])
def download_export_as(filename, fmt):
    """Download a history CSV file as CSV or XLSX."""
    if ".." in filename or "/" in filename:
        return "Invalid filename", 400
    if fmt not in ("csv", "xlsx"):
        return "不支援的格式", 400
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.isfile(path):
        return "檔案不存在", 404

    if filename.endswith(".csv") and fmt == "xlsx":
        rows = _read_csv_file(path)
        buf = _rows_to_xlsx(rows)
        dl_name = os.path.splitext(filename)[0] + ".xlsx"
        return send_file(
            buf, as_attachment=True, download_name=dl_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Default: serve the file as-is
    mimes = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".csv": "text/csv",
    }
    ext = os.path.splitext(filename)[1]
    mime = mimes.get(ext, "application/octet-stream")
    return send_file(path, as_attachment=True, download_name=filename, mimetype=mime)


@app.route("/exports/<filename>", methods=["DELETE"])
def delete_export(filename):
    if ".." in filename or "/" in filename:
        return "Invalid filename", 400
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.isfile(path):
        return "檔案不存在", 404
    os.remove(path)
    return jsonify({"ok": True})


@app.route("/exports/<filename>/rename", methods=["POST"])
def rename_export(filename):
    if ".." in filename or "/" in filename:
        return "Invalid filename", 400
    data = request.get_json(silent=True)
    if not data or not data.get("new_name"):
        return "Missing new_name", 400

    new_name = data["new_name"].strip()
    if ".." in new_name or "/" in new_name or "\\" in new_name:
        return "Invalid new name", 400

    old_ext = os.path.splitext(filename)[1]
    if not new_name.endswith((".csv", ".xlsx")):
        new_name += old_ext

    old_path = os.path.join(EXPORT_DIR, filename)
    new_path = os.path.join(EXPORT_DIR, new_name)
    if not os.path.isfile(old_path):
        return "檔案不存在", 404
    if os.path.exists(new_path) and old_path != new_path:
        return "檔名已存在", 409

    os.rename(old_path, new_path)
    return jsonify({"ok": True, "new_name": new_name})


@app.route("/import_csv", methods=["POST"])
def import_csv():
    """Import a previously exported CSV or XLSX to restore session data."""
    f = request.files.get("file")
    if not f or f.filename == "":
        return jsonify({"error": "缺少檔案"}), 400
    fname = f.filename.lower()
    if not fname.endswith((".csv", ".xlsx")):
        return jsonify({"error": "只支援 CSV / XLSX 檔案"}), 400

    try:
        if fname.endswith(".xlsx"):
            import tempfile
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            f.save(tmp.name)
            tmp.close()
            raw_rows = _read_xlsx_file(tmp.name)
            os.unlink(tmp.name)
        else:
            content = f.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            raw_rows = []
            for r in reader:
                raw_rows.append({
                    "工作項目編號": r.get("工作項目編號", ""),
                    "original_filename": r.get("原始檔名", ""),
                    "來文機關": r.get("來文機關", "N/A"),
                    "來文字號": r.get("來文字號", "N/A"),
                    "收文文號": r.get("收文文號", ""),
                    "收文日期": r.get("收文日期", ""),
                    "事由": r.get("事由", ""),
                    "status": r.get("status", "imported"),
                })

        rows = []
        for r in raw_rows:
            r["status"] = r.get("status") or "imported"
            r["_sort_key"] = ocrmod.parse_roc_date_to_sort_key(r.get("收文日期", ""))
            r["_file_hash"] = ""
            r["_image_file"] = ""
            rows.append(r)

        if not rows:
            return jsonify({"error": "檔案沒有資料"}), 400

        token = uuid.uuid4().hex[:12]
        _sort_and_renumber(rows)
        RESULTS[token] = rows

        safe_rows = [_safe_row(r) for r in rows]
        return jsonify({"token": token, "rows": safe_rows, "count": len(rows)})
    except Exception as e:
        return jsonify({"error": f"讀取檔案失敗：{e}"}), 400


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=os.environ.get("DEBUG") == "1")
